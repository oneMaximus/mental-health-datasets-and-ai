"""
Reddit Mental Health Data Collector
Collects posts from diagnosed users across multiple mental health subreddits
and exports to separate Excel sheets by disorder
"""

import praw
import pandas as pd
from datetime import datetime
import re
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Configuration
REDDIT_CONFIG = {
    'client_id': 'lB95YTUsFXu3iahQJm8wRw',
    'client_secret': 'sMMBGju4uPJSx7i8xcT9kfX_9dKC_g',
    'user_agent': 'mental_health_research_bot/1.0'
}

# Mental health conditions and their corresponding subreddits
DISORDERS = {
    'Depression': ['depression', 'depression_help', 'mentalhealth', 'depressed', 'depression_memes', 'depressionregimens'],
    'ADHD': ['ADHD', 'ADHDmemes', 'adhdwomen', 'adultadhd', 'adhd_anxiety', 'ADHDers'],
    'Anxiety': ['Anxiety', 'HealthAnxiety', 'socialanxiety', 'PanicAttack', 'Agoraphobia', 'Anxietyhelp'],
    'Bipolar': ['bipolar', 'bipolar2', 'BipolarReddit', 'BipolarSOs', 'bipolarart'],
    'PTSD': ['ptsd', 'CPTSD', 'traumatoolbox', 'PTSDCombat', 'rapecounseling'],
    'Autism': ['autism', 'aspergers', 'AutismTranslated', 'aspergirls', 'AutisticAdults', 'neurodiversity'],
    'Schizophrenia': ['schizophrenia', 'schizoaffective', 'Psychosis', 'Schizoaffective'],
    'OCD': ['OCD', 'OCDmemes', 'OCDRecovery', 'ROCD', 'PureO'],
    'Eating_Disorder': ['EatingDisorders', 'EDAnonymous', 'fuckeatingdisorders', 'AnorexiaNervosa', 'bulimia', 'BingeEatingDisorder']
}

# Keywords to identify diagnosed users
DIAGNOSIS_KEYWORDS = [
    r'\bdiagnosed\b',
    r'\bdiagnosis\b',
    r'\bformally diagnosed\b',
    r'\bofficially diagnosed\b',
    r'\bprofessionally diagnosed\b',
    r'\brecently diagnosed\b',
    r'\bjust diagnosed\b',
    r'\bmy diagnosis\b',
    r'\bgot diagnosed\b',
    r'\bwas diagnosed\b',
    r'\bmy doctor diagnosed\b',
    r'\bpsychiatrist diagnosed\b',
    r'\btherapist diagnosed\b',
    r'\bclinical diagnosis\b',
    r'\bmedical diagnosis\b'
]

def check_if_diagnosed(text):
    """
    Check if the post/comment mentions being diagnosed
    """
    if not text:
        return False
    
    text_lower = text.lower()
    
    # Check for diagnosis keywords
    for keyword in DIAGNOSIS_KEYWORDS:
        if re.search(keyword, text_lower):
            return True
    
    return False

def initialize_reddit():
    """
    Initialize Reddit API connection
    """
    try:
        reddit = praw.Reddit(
            client_id=REDDIT_CONFIG['client_id'],
            client_secret=REDDIT_CONFIG['client_secret'],
            user_agent=REDDIT_CONFIG['user_agent']
        )
        print("✓ Successfully connected to Reddit API")
        return reddit
    except Exception as e:
        print(f"✗ Error connecting to Reddit API: {e}")
        return None

def collect_posts_for_disorder(reddit, disorder_name, subreddit_list, target_posts=2000):
    """
    Collect posts from subreddits for a specific disorder
    Uses multiple sorting methods and date filtering to get historical data
    """
    posts_data = []
    seen_post_ids = set()  # Avoid duplicates
    
    # Date filter - posts from 2015 onwards
    cutoff_date = datetime(2015, 1, 1).timestamp()
    
    print(f"\n{'='*60}")
    print(f"Collecting data for: {disorder_name}")
    print(f"Target: {target_posts} posts from 2015 onwards")
    print(f"{'='*60}")
    
    for subreddit_name in subreddit_list:
        try:
            subreddit = reddit.subreddit(subreddit_name)
            print(f"\nSearching r/{subreddit_name}...")
            
            diagnosed_count = 0
            
            # Use multiple sorting methods to get diverse posts
            sorting_methods = [
                ('new', 1000),      # Recent posts
                ('top', 1000),      # Top posts of all time
                ('hot', 300),       # Currently hot posts
            ]
            
            for sort_method, limit in sorting_methods:
                print(f"  Fetching {sort_method} posts...")
                
                try:
                    if sort_method == 'new':
                        posts = subreddit.new(limit=limit)
                    elif sort_method == 'top':
                        posts = subreddit.top(time_filter='all', limit=limit)
                    elif sort_method == 'hot':
                        posts = subreddit.hot(limit=limit)
                    
                    for post in posts:
                        # Skip if already processed
                        if post.id in seen_post_ids:
                            continue
                        
                        # Check date - only posts from 2015 onwards
                        if post.created_utc < cutoff_date:
                            continue
                        
                        seen_post_ids.add(post.id)
                        
                        # Combine title and selftext for checking
                        full_text = f"{post.title} {post.selftext}"
                        
                        # Check if user mentions being diagnosed
                        if check_if_diagnosed(full_text):
                            diagnosed_count += 1
                            
                            # Extract post data
                            post_date = datetime.fromtimestamp(post.created_utc).strftime('%Y-%m-%d %H:%M:%S')
                            
                            posts_data.append({
                                'Date': post_date,
                                'Title': post.title,
                                'Description': post.selftext if post.selftext else '[No description]',
                                'Subreddit': f"r/{subreddit_name}",
                                'URL': f"https://reddit.com{post.permalink}",
                                'Score': post.score,
                                'Num_Comments': post.num_comments
                            })
                            
                            if diagnosed_count % 50 == 0:
                                print(f"    ✓ Found {diagnosed_count} diagnosed posts so far...")
                    
                    time.sleep(2)  # Rate limiting between sort methods
                    
                except Exception as e:
                    print(f"    ✗ Error with {sort_method} sorting: {e}")
                    continue
            
            print(f"  ✓ Completed r/{subreddit_name}: {diagnosed_count} diagnosed posts")
            
            # Check if we've reached target
            if len(posts_data) >= target_posts:
                print(f"\n✓ Reached target of {target_posts} posts!")
                break
            
        except Exception as e:
            print(f"  ✗ Error accessing r/{subreddit_name}: {e}")
            continue
        
        # Delay between subreddits
        time.sleep(3)
    
    print(f"\n✓ Total posts collected for {disorder_name}: {len(posts_data)}")
    return posts_data

def export_to_excel(all_data, output_file='mental_health_reddit_data.xlsx'):
    """
    Export collected data to Excel with separate sheets for each disorder
    """
    print(f"\n{'='*60}")
    print("Exporting data to Excel...")
    print(f"{'='*60}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for disorder_name, posts_data in all_data.items():
            if posts_data:
                # Create DataFrame
                df = pd.DataFrame(posts_data)
                
                # Reorder columns
                df = df[['Date', 'Title', 'Description', 'Subreddit', 'Score', 'Num_Comments', 'URL']]
                
                # Sort by date (newest first)
                df['Date'] = pd.to_datetime(df['Date'])
                df = df.sort_values('Date', ascending=False)
                df['Date'] = df['Date'].dt.strftime('%Y-%m-%d %H:%M:%S')
                
                # Write to Excel
                sheet_name = disorder_name[:31]  # Excel sheet name limit
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                print(f"✓ Created sheet: {sheet_name} ({len(posts_data)} posts)")
            else:
                print(f"✗ No data for {disorder_name}, skipping sheet")
    
    # Format Excel file
    format_excel_file(output_file)
    
    print(f"\n✓ Data exported successfully to: {output_file}")

def format_excel_file(filename):
    """
    Format the Excel file for better readability
    """
    wb = load_workbook(filename)
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # Date
        ws.column_dimensions['B'].width = 50  # Title
        ws.column_dimensions['C'].width = 80  # Description
        ws.column_dimensions['D'].width = 20  # Subreddit
        ws.column_dimensions['E'].width = 12  # Score
        ws.column_dimensions['F'].width = 15  # Num_Comments
        ws.column_dimensions['G'].width = 60  # URL
        
        # Wrap text in Description column
        for row in ws.iter_rows(min_row=2, max_col=7):
            row[2].alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(filename)

def print_summary(all_data):
    """
    Print summary statistics
    """
    print(f"\n{'='*60}")
    print("COLLECTION SUMMARY")
    print(f"{'='*60}")
    
    total_posts = 0
    for disorder_name, posts_data in all_data.items():
        count = len(posts_data)
        total_posts += count
        print(f"{disorder_name:20}: {count:4} posts")
    
    print(f"{'-'*60}")
    print(f"{'TOTAL':20}: {total_posts:4} posts")
    print(f"{'='*60}")

def main():
    """
    Main execution function
    """
    print("="*60)
    print("REDDIT MENTAL HEALTH DATA COLLECTOR")
    print("Collecting posts from DIAGNOSED users only")
    print("Target: ~2000 posts per disorder (2015-present)")
    print("="*60)
    
    # Initialize Reddit connection
    reddit = initialize_reddit()
    if not reddit:
        print("\n✗ Failed to connect to Reddit. Please check your credentials.")
        return
    
    # Collect data for all disorders
    all_data = {}
    
    for disorder_name, subreddit_list in DISORDERS.items():
        posts_data = collect_posts_for_disorder(
            reddit, 
            disorder_name, 
            subreddit_list,
            target_posts=2000  # Target 2000 posts per disorder
        )
        all_data[disorder_name] = posts_data
    
    # Export to Excel
    if any(all_data.values()):
        output_file = f'mental_health_reddit_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        export_to_excel(all_data, output_file)
        print_summary(all_data)
    else:
        print("\n✗ No data collected. Please check your Reddit API connection and subreddit access.")

if __name__ == "__main__":
    main()